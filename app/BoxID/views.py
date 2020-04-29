from django.shortcuts import render
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from .models import ModelStage,SerialNumberInfo,Box
# from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from app.access_control import access_control
from datetime import datetime,timedelta,date
# import datetime
from django.conf import settings
import random
import string
import os
import time
from openpyxl import load_workbook,Workbook
import json
import xlrd
from bs4 import BeautifulSoup
import filetype
import requests
from lxml import etree,html
from django.db.models import Q,F
import copy
from django.db.models import Max
#信息导入表中的数据(sn和modelstage)
def Import_models(request):#这里需要考虑到是否设备料名有变更，如果变更的话，就需要把绑定解除
    if request.method == "POST":
        try:
            UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            model = request.POST['import_Models']
            db_data = list(ModelStage.objects.filter(ModelStage=model).values()) #查找DB里面是否含有数据
            scrapy_data = Scrapy_ModelstageId(model)
            #数据库里面没有查找到这个数据 意思是创建新的
            if len(db_data) ==0 :
                if len(scrapy_data) !=0:
                    for items in scrapy_data:
                        ModelStage.objects.create(ModelStage=items["ModelStage"],PartCode=items["PartCode"],
                                                  PartName=items["PartName"],RequireQuantity=items["RequireQuantity"],
                                                  UpdatedTime=UpdatedTime)
                    return restful.ok(data=scrapy_data,message="data import success")
                else:
                    return restful.params_error(message="AEMS no this "+model)
            #表数据库里面查找到了数据
            else:
                add_data = copy.deepcopy(scrapy_data)
                dbb_data = copy.deepcopy(db_data)
                for i in db_data:
                    for j in scrapy_data:
                        if i["PartCode"] == j["PartCode"] and i["RequireQuantity"] == j["RequireQuantity"]:
                            dbb_data.remove(i)
                            add_data.remove(j)
                if len(dbb_data) ==0 and len(add_data) ==0:          #表示没有必要更新的数据属于重复输入
                    return restful.params_error(message="models no need update")
                if len(dbb_data) !=0 or len(add_data) !=0:
                    for m in dbb_data:
                        ModelStage.objects.filter(ModelStage=m["ModelStage"],PartCode=m["PartCode"]).delete()
                    for n in add_data:
                        ModelStage.objects.create(ModelStage=n["ModelStage"], PartCode=n["PartCode"],
                                                  PartName=n["PartName"], RequireQuantity=n["RequireQuantity"],
                                                  UpdatedTime=UpdatedTime)
                    message = Box.objects.filter(ModelStageId=model).values_list("BoxId",flat=True).distinct()
                    show_message = ""
                    if len(message)!=0:
                        show_message =",".join(message)+" Need to Check"

                    return restful.ok(data=add_data, message="update OK "+show_message)
        except:
            return restful.params_error(message="no this models")
#导入SN信息
def Import_SN(request):
    if request.method == "POST":
        try:
            UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            SN = request.POST['import_SNs']
            if len(SerialNumberInfo.objects.filter(SerialNumber=SN).values("Id")) == 0:
                scrapy_sn = Scrapy_SN(SN)
                if len(scrapy_sn) != 0:
                    for item in scrapy_sn:
                        SerialNumberInfo.objects.create(SerialNumber=item["BarCode"],PartCode=item["PartCode"],UpdatedTime=UpdatedTime)
                    return restful.ok(data=scrapy_sn[0],message="data import success")
                else:
                    return restful.params_error(message="AEMS no this "+SN)
            else:
                return restful.params_error(message=SN+" had exist in AEMSLite system")
        except:
            return restful.params_error(message="error")
#获取box的最新的那个
# def Get_Box(request):
#     if request.method == "POST":
#         try:
#             data = list(Box.objects.all().order_by("-Id").values("BoxId"))[0]
#             if len(data)==0:
#                 data1={"BoxId":"BOX0000000"}
#                 return restful.ok(data=data1)
#             else:
#                 da = int(data["BoxId"][3:])+1
#                 num = '0'
#                 if len(str(da))<7:
#                     add_zero = 7-len(str(da))
#                     for i in range(0,add_zero-1):
#                         num +='0'
#                 data_str = "BOX"+num+str(da)
#
#                 dt = {"BoxId": data_str}
#                 return restful.ok(data=dt)
#         except:
#             return restful.params_error(message="error")
#增加盒子
def Add_Box(request):
    if request.method == "POST":
        try:
            box_id = request.POST['import_Box']
            check_box = list(Box.objects.filter(BoxId=box_id).values())
            UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if len(check_box) !=0:
                return restful.params_error(message=box_id+" had exist")
            else:
                Box.objects.create(BoxId=box_id,UpdatedTime=UpdatedTime)
                return restful.ok(message=box_id+" create successful")
        except:
            return restful.params_error(message="error")


#绑定页面的数据从数据表中去获取要绑定的SN和modelstage
def ModelstageID(request):
    if request.method == "POST":
        try:
            model = request.POST['modelstageId']
            data = list(ModelStage.objects.filter(ModelStage=model).order_by("Id").values())
            if len(data) != 0:
                for i in data:
                    i["BrushSN"]=""
                return restful.ok(data=data)
            else:
                return restful.params_error(message=model+" not exist")
        except:
            return restful.params_error(message="error")
def Box_Id(request):
    if request.method == "POST":
        try:
            box_id = request.POST['BoxId']
            data1 = list(Box.objects.filter(BoxId=box_id).values())
            if len(data1) ==0:
                return restful.params_error(message="AEMSLite system had't this "+box_id)
            elif data1[0]["ModelStageId"] !=None:
                #说明这里有绑定的事物 返回绑定的数据
                data = MapBoxModelstage(box_id)
                brushsn = []
                for i in data:
                    i["BoxID"] = box_id
                    if i["BrushSN"] != "":
                        brushsn.extend(i["BrushSN"].split(","))
                check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=data1[0]["Id"]).exclude(
                    SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                Find_no = {}
                if len(check_no_find) != 0:
                    Find_no["Id"] = ""
                    Find_no["BoxID"] = box_id
                    Find_no["ModelStage"] = ""
                    Find_no["PartCode"] = ""
                    Find_no["PartName"] = ""
                    Find_no["RequireQuantity"] = ""
                    Find_no["BrushNumber"] = len(check_no_find)
                    Find_no["BrushSN"] = ",".join(check_no_find)
                    Find_no["AlternativePartCode"] = ""
                    Find_no["Status"] = "none"
                    data.append(Find_no)

                return restful.ok(data=data,message="BoxId had binding")
            else:
                return restful.ok(message=box_id+" add BOXId Correct")
        except:
            return restful.params_error(message='error')
#绑定业务的实现功能
def SNinfo(request):
    if request.method == "POST":
        try:
            sn = request.POST['SN']
            models = request.POST['ModelStage']
            BoxId = request.POST['BoxId']
            sn_info = list(SerialNumberInfo.objects.filter(SerialNumber=sn).values("Id","PartCode","BoxIdBindingId"))#查询是否在DB里面存在 两种 存在和不存在

            check_box = list(Box.objects.filter(BoxId=BoxId).values())
            check_modelstage = list(ModelStage.objects.filter(ModelStage=models).values())
            # 验证boxId是否存在
            if len(check_box) == 0:
                return restful.params_error(message=BoxId + " input message is error")
            # 验证models是否存在
            if len(check_modelstage) == 0:
                return restful.params_error(message=models + " input message is error")
            # 验证box的绑定的modelstageId是否是该model，如果不是就要换box来用
            if check_box[0]["ModelStageId"] !=None and check_box[0]["ModelStageId"] != check_modelstage[0]["ModelStage"]:
                return restful.params_error(message="This "+BoxId +" had binding other ModelstageId,please change other box")

            if len(sn_info) != 0:              #表示DB里面存在
                #验证第一步，该SN是否有绑定box
                Check_Box_bing = list(Box.objects.filter(Id=sn_info[0]["BoxIdBindingId"]).values())
                if len(Check_Box_bing) !=0:
                    return restful.params_error(message=sn + " had binding")
                #验证第二步 是否在该modelstage里面的料号里面并且相同
                check_SN= list(ModelStage.objects.filter(ModelStage=models).filter(Q(PartCode=sn_info[0]["PartCode"])|Q(AlternativePartCode__contains=sn_info[0]["PartCode"]))
                               .values("ModelStage","RequireQuantity", "AlternativePartCode","PartCode"))   #确认是否存在 代用料或者使用料中
                if len(check_SN) !=0:
                    #卡数量
                    if check_SN[0]["AlternativePartCode"] !=None:
                        ParctCodes = tuple(check_SN[0]["PartCode"]+check_SN[0]["AlternativePartCode"].split(","))
                    else:
                        ParctCodes = tuple(check_SN[0]["PartCode"])
                    check_SNNumber =SerialNumberInfo.objects.filter(BoxIdBindingId=check_box[0]["Id"],PartCode__in=ParctCodes).count()
                    if check_SNNumber >= check_SN[0]["RequireQuantity"]:
                        return restful.params_error(message=sn + " " + sn_info[0]["PartCode"] + " qty overload")
                    else:
                        SerialNumberInfo.objects.filter(SerialNumber=sn).update(BoxIdBindingId=check_box[0]["Id"])  # 更新sn的绑定的盒子id
                        Box.objects.filter(Id=check_box[0]["Id"]).update(ModelStageId=check_SN[0]["ModelStage"])  # 更新盒子的ModelstageId字段
                        data = MapBoxModelstage(check_box[0]["BoxId"])
                        brushsn = []
                        for i in data:
                            i["BoxID"] = BoxId
                            if i["BrushSN"] != "":
                                brushsn.extend(i["BrushSN"].split(","))
                        check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_box[0]["Id"]).exclude(
                            SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                        Find_no = {}
                        if len(check_no_find) != 0:
                            Find_no["Id"] = ""
                            Find_no["BoxID"] = BoxId
                            Find_no["ModelStage"] = ""
                            Find_no["PartCode"] = ""
                            Find_no["PartName"] = ""
                            Find_no["RequireQuantity"] = ""
                            Find_no["BrushNumber"] = len(check_no_find)
                            Find_no["BrushSN"] = ",".join(check_no_find)
                            Find_no["AlternativePartCode"] = ""
                            Find_no["Status"] = "none"
                            data.append(Find_no)
                        return restful.ok(data=data, message=sn + " binding finished")
                else:#既不在代用料和使用料中
                    return restful.params_error(message="SN not in this "+models)
            else:#表示该SN不在DB里面，而且没有绑定数据
                UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                data = Scrapy_SN(sn)
                if len(data) !=0:
                    # 验证只需要是否在该modelstage里面的料号里面并且相同
                    check_SN = list(ModelStage.objects.filter(ModelStage=models).filter(
                        Q(PartCode=data[0]["PartCode"]) | Q(AlternativePartCode__contains=data[0]["PartCode"]))
                                    .values("ModelStage", "RequireQuantity", "AlternativePartCode","PartCode"))  # 确认是否存在 代用料或者使用料中
                    if len(check_SN) != 0:
                        # 卡数量
                        if check_SN[0]["AlternativePartCode"] != None:
                            ParctCodes = tuple(check_SN[0]["PartCode"] + check_SN[0]["AlternativePartCode"].split(","))
                        else:
                            ParctCodes = tuple(check_SN[0]["PartCode"])
                        # ParctCodes = tuple(check_SN[0]["PartCode"] + check_SN[0]["AlternativePartCode"].split(","))
                        check_SNNumber = SerialNumberInfo.objects.filter(BoxIdBindingId=check_box[0]["Id"],
                                                                         PartCode__in=ParctCodes).count()
                        if check_SNNumber >= check_SN[0]["RequireQuantity"]:
                            return restful.params_error(message=sn + " " + sn_info[0]["PartCode"] + " qty overload")
                        else:
                            # SerialNumberInfo.objects.filter(SerialNumber=sn).update(BoxIdBindingId=check_box[0]["Id"])  # 更新sn的绑定的盒子id
                            SerialNumberInfo.objects.create(SerialNumber=data[0]["BarCode"],PartCode=data[0]["PartCode"], UpdatedTime=UpdatedTime,BoxIdBindingId=check_box[0]["Id"])
                            Box.objects.filter(Id=check_box[0]["Id"]).update( ModelStageId=check_SN[0]["ModelStage"])  # 更新盒子的ModelstageId字段
                            data = MapBoxModelstage(check_box[0]["BoxId"])
                            brushsn = []
                            for i in data:
                                i["BoxID"] = BoxId
                                if i["BrushSN"] != "":
                                    brushsn.extend(i["BrushSN"].split(","))
                            check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_box[0]["Id"]).exclude(
                                SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                            Find_no = {}
                            if len(check_no_find) != 0:
                                Find_no["Id"] = ""
                                Find_no["BoxID"] = BoxId
                                Find_no["ModelStage"] = ""
                                Find_no["PartCode"] = ""
                                Find_no["PartName"] = ""
                                Find_no["RequireQuantity"] = ""
                                Find_no["BrushNumber"] = len(check_no_find)
                                Find_no["BrushSN"] = ",".join(check_no_find)
                                Find_no["AlternativePartCode"] = ""
                                Find_no["Status"] = "none"
                                data.append(Find_no)
                            return restful.ok(data=data, message=sn + " binding finished")
                    else:  # 既不在代用料和使用料中
                        return restful.params_error(message="SN not in this " + models)
                else:
                    return restful.params_error(message="get AEMS system no this " + sn)
        except Exception as e:
            return restful.params_error(message=e)



#解绑业务的查询数据modelstageId和BoxId还有SN信息的查询
def unbinding_query(request):
    if request.method == "POST":
        try:
            BoxId = request.POST['BoxId']
            SN = request.POST['SN']
            #查询boxId,SN的存在性
            check_SN = list(SerialNumberInfo.objects.filter(SerialNumber=SN).values())
            check_boxid = list(Box.objects.filter(BoxId=BoxId).values())
            #查询为空的时候返回信息为空 查无此信息
            if len(check_SN)==0 and len(check_boxid)==0:
                return restful.params_error(message="query info is empty")

            if len(check_boxid)!=0 and len(check_SN) ==0:
                data = MapBoxModelstage(BoxId)
                if len(data)==0:
                    return restful.params_error(message=BoxId+" no binding Modelstage")
                brushsn =[]
                for i in data:
                    i["BoxID"] = BoxId
                    if i["BrushSN"] !="":
                        brushsn.extend(i["BrushSN"].split(","))
                check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).exclude(SerialNumber__in=tuple(brushsn)).values_list("SerialNumber",flat=True)
                Find_no={}
                if len(check_no_find) !=0:
                    Find_no["Id"]=""
                    Find_no["BoxID"]=BoxId
                    Find_no["ModelStage"]=""
                    Find_no["PartCode"]=""
                    Find_no["PartName"]=""
                    Find_no["RequireQuantity"]=""
                    Find_no["BrushNumber"]=len(check_no_find)
                    Find_no["BrushSN"]= ",".join(check_no_find)
                    Find_no["AlternativePartCode"]= ""
                    Find_no["Status"]= "none"
                    data.append(Find_no)
                return restful.ok(data=data)
            if len(check_SN) !=0 and check_SN[0]["BoxIdBindingId"] !=None:
                box = list(Box.objects.filter(Id=check_SN[0]["BoxIdBindingId"]).values())
                sns = SerialNumberInfo.objects.filter(BoxIdBindingId=check_SN[0]["BoxIdBindingId"],PartCode=check_SN[0]["PartCode"]).count()
                sn = SerialNumberInfo.objects.filter(BoxIdBindingId=check_SN[0]["BoxIdBindingId"],PartCode=check_SN[0]["PartCode"]).values_list("SerialNumber",flat=True)
                data = list(ModelStage.objects.filter(ModelStage=box[0]["ModelStageId"]).filter(Q(PartCode=check_SN[0]["PartCode"])|Q(AlternativePartCode__contains=check_SN[0]["PartCode"])).values())
                if len(data) !=0:
                    for i in data:
                        i["BrushSN"]=",".join(sn)
                        i["BrushNumber"] = sns
                        i["BoxID"] = box[0]["BoxId"]
                        if i["RequireQuantity"] ==sns:
                            i["Status"]="OK"
                        if i["RequireQuantity"] < sns:
                            i["Status"] = "Over"
                    return restful.ok(data=data)
                else:
                    Find_no = {}
                    Find_no["Id"] = ""
                    Find_no["BoxID"] = box[0]["BoxId"]
                    Find_no["ModelStage"] = ""
                    Find_no["PartCode"] = ""
                    Find_no["PartName"] = ""
                    Find_no["RequireQuantity"] = ""
                    Find_no["BrushNumber"] = 1
                    Find_no["BrushSN"] = SN
                    Find_no["AlternativePartCode"] = ""
                    Find_no["Status"] = "none"
                    return restful.ok(data=[Find_no])
        except Exception as e:
            return restful.params_error(message=e)

#单个的解绑任务的完成
def release_PartCode(request):
    try:
        if request.method == "POST":
            id = request.POST['id']
            release_brushNumber = int(request.POST['release_brushNumber'])
            check_Bush_SN = request.POST['check_Bush_SN']
            release_partCode = request.POST['release_partCode']
            if release_partCode not in check_Bush_SN:
                return restful.params_error(message= release_partCode+" not in this"+check_Bush_SN)
            else:
                sns = list(SerialNumberInfo.objects.filter(SerialNumber=release_partCode).values())
                binding = SerialNumberInfo.objects.filter(BoxIdBindingId=sns[0]["BoxIdBindingId"]).count()
                if binding == 1:
                    Box.objects.filter(Id=sns[0]["BoxIdBindingId"]).update(ModelStageId=None)
                SerialNumberInfo.objects.filter(SerialNumber=release_partCode).update(BoxIdBindingId=None)
                return restful.ok(message="release success")
    except Exception as e:
        return restful.params_error(message=e)
#批量的解绑动作的实现
def ReleaseAllPartCode(request):
    try:
        if request.method == "POST":
            models_bindingSN = request.POST.get('models_ids')
            #查询SN的信息
            if models_bindingSN !="":
                sns = models_bindingSN.split(",")
                Sn_obj = list(SerialNumberInfo.objects.filter(SerialNumber__in=tuple(sns)).values("BoxIdBindingId"))
                Sn_count = SerialNumberInfo.objects.filter(BoxIdBindingId=Sn_obj[0]["BoxIdBindingId"]).count()
                if Sn_count == len(sns):
                    Box.objects.filter(Id=Sn_obj[0]["BoxIdBindingId"]).update(ModelStageId=None)
                SerialNumberInfo.objects.filter(SerialNumber__in=tuple(sns)).update(BoxIdBindingId=None)
                return restful.ok(message="Batch operation successful")
            else:
                return restful.params_error(message="submit SN no need unbinding")
    except Exception as e:
        return restful.params_error(message=e)



#删除动作的实现
def remove_PartCode(request):
    try:
        if request.method == "POST":
            id = request.POST['id']
            check_data = list(ModelStage.objects.filter(Id=id).values("BrushNumber"))
            if check_data[0]['BrushNumber'] ==0:
                ModelStage.objects.filter(Id=id).delete()
                return restful.ok(message="delete success")
            else:
                return restful.params_error(message="need first to release")
    except Exception as e:
        return restful.params_error(message=e)
#单个的进行数量的更改
def modify_PartCode(request):
    try:
        if request.method == "POST":
            id = request.POST['id']
            modify_requirement = request.POST['modify_requirement']
            check_data = list(ModelStage.objects.filter(Id=id).values("BrushNumber","RequireQuantity"))
            if int(modify_requirement) > check_data[0]["BrushNumber"]:
                ModelStage.objects.filter(Id=id).update(RequireQuantity=modify_requirement,Status="Open")
                return restful.ok(message="modify requirement success")
            if int(modify_requirement) < check_data[0]["BrushNumber"]:
                return restful.params_error(message="need first to release this model")
            if int(modify_requirement) == check_data[0]["BrushNumber"]:
                return restful.ok(message="no need to modify")
    except Exception as e:
        return restful.params_error(message=e)




#查询需要更改代用料的models
def Alternative_query(request):
    if request.method == "POST":
        try:
            PartCode = request.POST['PartCode']
            ModelStage = request.POST['ModelStage']
            sql = 'SELECT "Id","ModelStage","PartCode","PartName","RequireQuantity","AlternativeType","AlternativePartCode" FROM "ModelStage" where 1=1 '
            if len(ModelStage) !=0:
                sql = sql+'and "ModelStage"= \'' + ModelStage + '\''
            if PartCode !="":
                sql = sql +' and "PartCode"=\'' + PartCode + '\''
            cur = connection.cursor()
            cur.execute(sql, None)
            desc = cur.description
            data =[dict(zip([col[0] for col in desc], row)) for row in cur.fetchall()]
            if len(data) !=0:
                return restful.ok(data=data)
            else:
                return restful.params_error(message=" query is empty")
        except:
            return restful.params_error(message="error")
#修改代用料
def Modify_alterParCode(request):
    if request.method=="POST":
        try:
            model_id = request.POST['model_id']
            original_parcode = request.POST['original_parcode']
            add_ParCode = request.POST['add_ParCode']
            alter_type = request.POST['alter_type']
            #检查有没有重复的料号
            check_add = list(ModelStage.objects.filter(PartCode=add_ParCode).values("Id"))
            if len(check_add) !=0:
                return restful.params_error(message=add_ParCode+' had used')

            if len(add_ParCode) !=0:
                #检查是否在其中
                if add_ParCode in original_parcode.split(','):
                    return restful.params_error(message=add_ParCode +' had exsit '+original_parcode)
                else:
                    update_parcode = original_parcode+add_ParCode+','
                    ModelStage.objects.filter(Id=model_id).update(AlternativePartCode=update_parcode,AlternativeType=int(alter_type))
                    return restful.ok(message=" add success")
            else:
                return restful.params_error(message="submit ParCode is error")
        except Exception as e:
            return restful.params_error(message=e)
#删除代用料
def Delete_alterParCode(request):
    if request.method == "POST":
        try:
            model_id = request.POST['model_id']
            original_parcode = request.POST['original_parcode']
            del_ParCode = request.POST['del_ParCode']
            #这里需要去检查是否有绑定，如果有绑定的话，就需要去解绑
            check_SN = list(SerialNumberInfo.objects.filter(PartCode=del_ParCode).values("BoxIdBindingId"))
            if len(check_SN) !=0:
                boxname = list(Box.objects.filter(Id=check_SN[0]["BoxIdBindingId"]).values())[0]["BoxId"]
                return restful.params_error(message=boxname+" this "+del_ParCode+" need to release")
            # 检查是否有该料号
            if len(original_parcode) !=0:
                pc = original_parcode.split(",")
                if del_ParCode not in pc:
                    return restful.params_error(original_parcode+" have no this "+del_ParCode)
                else:
                    str = ""
                    for i in pc:
                        if i==del_ParCode or i=='':
                            pc.remove(i)
                    for j in pc:
                        str += j+','
                    ModelStage.objects.filter(Id=model_id).update(AlternativePartCode=str)
                    return restful.ok(message=del_ParCode+' delete success')
            else:
                return restful.params_error(message=original_parcode+" have no this "+del_ParCode)
        except Exception as e:
            return restful.params_error(message=e)


#查询所有的信息数据
def Info_all_query(request):
    # try:
        BoxId = request.POST['BoxId']
        ModelStageID = request.POST['ModelStageID']
        SN = request.POST['SN']
        check_box = request.POST['check_box']
        # 查询boxId, ModestageId ，SN的存在性
        check_SN = list(SerialNumberInfo.objects.filter(SerialNumber=SN).values())
        check_boxid = list(Box.objects.filter(BoxId=BoxId).values())
        check_Model = list(ModelStage.objects.filter(ModelStage=ModelStageID).values())
        # 查询为空的时候返回信息为空 查无此信息
        if len(check_SN) == 0 and len(check_boxid) == 0 and len(check_Model) == 0:
            return restful.params_error(message="query info is empty")
        #表明是使用查询所有的数据表形式
        if check_box == 'false':

            if len(check_SN) !=0:
                data = MapSNModelstage(SN)
                if len(data) !=0:
                    return restful.ok(data=data)
                else:
                    return restful.params_error(message="query info is empty")
            if len(check_SN) ==0 and len(check_boxid) !=0 and len(check_Model) !=0:
                # check models 绑定的box是否是它 不是就返回空数据
                if check_boxid[0]["ModelStageId"] !=check_Model[0]["ModelStage"]:
                    return restful.params_error(message="query info is empty")
                else:
                    data = MapBoxModelstage(BoxId)
                    if len(data) == 0:
                        return restful.params_error(message="query info is empty")
                    brushsn = []
                    for i in data:
                        i["BoxID"] = BoxId
                        if i["BrushSN"] != "":
                            brushsn.extend(i["BrushSN"].split(","))
                    check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).exclude(
                        SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                    Find_no = {}
                    if len(check_no_find) != 0:
                        Find_no["Id"] = ""
                        Find_no["BoxID"] = BoxId
                        Find_no["ModelStage"] = ""
                        Find_no["PartCode"] = ""
                        Find_no["PartName"] = ""
                        Find_no["RequireQuantity"] = ""
                        Find_no["BrushNumber"] = len(check_no_find)
                        Find_no["BrushSN"] = ",".join(check_no_find)
                        Find_no["AlternativePartCode"] = ""
                        Find_no["AlternativeType"] = ""
                        Find_no["Status"] = "none"
                        data.append(Find_no)
                    return restful.ok(data=data)

            if len(check_SN) ==0 and len(check_boxid) !=0 and len(check_Model) ==0:
                data = MapBoxModelstage(BoxId)
                if len(data) == 0:
                    return restful.params_error(message=BoxId + " no binding Modelstage")
                brushsn = []
                for i in data:
                    i["BoxID"] = BoxId
                    if i["BrushSN"] != "":
                        brushsn.extend(i["BrushSN"].split(","))
                check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).exclude(
                    SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                Find_no = {}
                if len(check_no_find) != 0:
                    Find_no["Id"] = ""
                    Find_no["BoxID"] = BoxId
                    Find_no["ModelStage"] = ""
                    Find_no["PartCode"] = ""
                    Find_no["PartName"] = ""
                    Find_no["RequireQuantity"] = ""
                    Find_no["BrushNumber"] = len(check_no_find)
                    Find_no["BrushSN"] = ",".join(check_no_find)
                    Find_no["AlternativePartCode"] = ""
                    Find_no["AlternativeType"] = ""
                    Find_no["Status"] = "none"
                    data.append(Find_no)
                return restful.ok(data=data)
            if len(check_SN) ==0 and len(check_boxid) ==0 and len(check_Model) !=0:
                boxnamelist = Box.objects.filter(ModelStageId=check_Model[0]["ModelStage"]).values_list("BoxId",flat=True)
                html_data =[]
                for j in boxnamelist:
                    data = MapBoxModelstage(j)
                    check_boxname = list(Box.objects.filter(BoxId=j).values())
                    brushsn = []
                    for i in data:
                        i["BoxID"] = j
                        if i["BrushSN"] != "":
                            brushsn.extend(i["BrushSN"].split(","))
                    check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxname[0]["Id"]).exclude(
                        SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                    Find_no = {}
                    if len(check_no_find) != 0:
                        Find_no["Id"] = ""
                        Find_no["BoxID"] = BoxId
                        Find_no["ModelStage"] = ""
                        Find_no["PartCode"] = ""
                        Find_no["PartName"] = ""
                        Find_no["RequireQuantity"] = ""
                        Find_no["BrushNumber"] = len(check_no_find)
                        Find_no["BrushSN"] = ",".join(check_no_find)
                        Find_no["AlternativePartCode"] = ""
                        Find_no["AlternativeType"] = ""
                        Find_no["Status"] = "none"
                        data.append(Find_no)
                    html_data.extend(data)
                return restful.ok(data=html_data)
        # 表明是使用查询挂的东西
        else:
            dispaly_BOX = Box.objects.filter()
            if len(check_boxid) !=0:
                dispaly_BOX=dispaly_BOX.filter(BoxId=BoxId)
            if len(check_Model) !=0:
                dispaly_BOX=dispaly_BOX.filter(ModelStageId__contains=ModelStageID)
            if len(check_SN) !=0:
                dispaly_BOX = dispaly_BOX.filter(Id=check_SN[0]["BoxIdBindingId"])
            data = list(dispaly_BOX.values("BoxId","ModelStageId"))
            return restful.ok(data=data)

    # except Exception as e:
    #     return restful.params_error(message=e)
#查询之后的数据进行下载功能的实现
def Download_binding_infos(request):
    try:
        if request.method =="POST":
            BoxId = request.POST['BoxId']
            ModelStageID = request.POST['ModelStageID']
            SN = request.POST['SN']
            tb1 = ['BoxID', 'ModelStage', '料号', '品名', '需求数量', '已刷数量', '已刷SN', '代用(1)/混用(2)', '代用/混用料号','状态']
            check_SN = list(SerialNumberInfo.objects.filter(SerialNumber=SN).values())
            check_boxid = list(Box.objects.filter(BoxId=BoxId).values())
            check_Model = list(ModelStage.objects.filter(ModelStage=ModelStageID).values())
            # 查询为空的时候返回信息为空 查无此信息
            query_data=[]
            if len(check_SN) == 0 and len(check_boxid) == 0 and len(check_Model) == 0:
                query_data=[]
            if len(check_SN) !=0:
                data = MapSNModelstage(SN)
                if len(data) !=0:
                    query_data.extend(data)
                else:
                    query_data=[]
            if len(check_SN) ==0 and len(check_boxid) !=0 and len(check_Model) !=0:
                # check models 绑定的box是否是它 不是就返回空数据
                if check_boxid[0]["ModelStageId"] !=check_Model[0]["ModelStage"]:
                    query_data = []
                else:
                    data = MapBoxModelstage(BoxId)
                    if len(data) == 0:
                        query_data = []
                    brushsn = []
                    for i in data:
                        i["BoxID"] = BoxId
                        if i["BrushSN"] != "":
                            brushsn.extend(i["BrushSN"].split(","))
                    check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).exclude(
                        SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                    Find_no = {}
                    if len(check_no_find) != 0:
                        Find_no["Id"] = ""
                        Find_no["BoxID"] = BoxId
                        Find_no["ModelStage"] = ""
                        Find_no["PartCode"] = ""
                        Find_no["PartName"] = ""
                        Find_no["RequireQuantity"] = ""
                        Find_no["BrushNumber"] = len(check_no_find)
                        Find_no["BrushSN"] = ",".join(check_no_find)
                        Find_no["AlternativePartCode"] = ""
                        Find_no["AlternativeType"] = ""
                        Find_no["Status"] = "none"
                        data.append(Find_no)
                    query_data.extend(data)

            if len(check_SN) ==0 and len(check_boxid) !=0 and len(check_Model) ==0:
                data = MapBoxModelstage(BoxId)
                if len(data) == 0:
                    query_data=[]
                brushsn = []
                for i in data:
                    i["BoxID"] = BoxId
                    if i["BrushSN"] != "":
                        brushsn.extend(i["BrushSN"].split(","))
                check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).exclude(
                    SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                Find_no = {}
                if len(check_no_find) != 0:
                    Find_no["Id"] = ""
                    Find_no["BoxID"] = BoxId
                    Find_no["ModelStage"] = ""
                    Find_no["PartCode"] = ""
                    Find_no["PartName"] = ""
                    Find_no["RequireQuantity"] = ""
                    Find_no["BrushNumber"] = len(check_no_find)
                    Find_no["BrushSN"] = ",".join(check_no_find)
                    Find_no["AlternativePartCode"] = ""
                    Find_no["AlternativeType"] = ""
                    Find_no["Status"] = "none"
                    data.append(Find_no)
                query_data.extend(data)
            if len(check_SN) ==0 and len(check_boxid) ==0 and len(check_Model) !=0:
                boxnamelist = Box.objects.filter(ModelStageId=check_Model[0]["ModelStage"]).values_list("BoxId",flat=True)
                html_data =[]
                for j in boxnamelist:
                    data = MapBoxModelstage(j)
                    check_boxname = list(Box.objects.filter(BoxId=j).values())
                    brushsn = []
                    for i in data:
                        i["BoxID"] = j
                        if i["BrushSN"] != "":
                            brushsn.extend(i["BrushSN"].split(","))
                    check_no_find = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxname[0]["Id"]).exclude(
                        SerialNumber__in=tuple(brushsn)).values_list("SerialNumber", flat=True)
                    Find_no = {}
                    if len(check_no_find) != 0:
                        Find_no["Id"] = ""
                        Find_no["BoxID"] = BoxId
                        Find_no["ModelStage"] = ""
                        Find_no["PartCode"] = ""
                        Find_no["PartName"] = ""
                        Find_no["RequireQuantity"] = ""
                        Find_no["BrushNumber"] = len(check_no_find)
                        Find_no["BrushSN"] = ",".join(check_no_find)
                        Find_no["AlternativePartCode"] = ""
                        Find_no["AlternativeType"] = ""
                        Find_no["Status"] = "none"
                        data.append(Find_no)
                    html_data.extend(data)
                query_data.extend(html_data)

            #把数据具体换
            data_sheet =[]
            data_sheet.append(tb1)
            if len(query_data) !=0:
                for i in query_data:
                    data_sheet.append([i["BoxID"],i["ModelStage"],i["PartCode"],i["PartName"],i["RequireQuantity"],i["BrushNumber"],i["BrushSN"],i["AlternativeType"],i["AlternativePartCode"],i["Status"]])

                time_num = int(time.time())
                time_num = str(time_num)
                filename = 'binding_info_' + time_num + '.xlsx'
                wb = Workbook()
                index = 0
                sheet_name = "绑定信息报表"
                wb.create_sheet(sheet_name, index=index)
                sheet = wb[sheet_name]
                for row in data_sheet:
                    sheet.append(row)

                wb.save(os.path.join(settings.MEDIA_CHANGE_ROOT, filename))
                file_url = request.build_absolute_uri(settings.MEDIA_CHANGE_URL + filename)
                data = [file_url]
                return restful.ok(data=data)
            else:
                return restful.params_error(message="query is empty")
    except Exception as e:
        return restful.params_error(message=e)

#这里是进行Box去map物料的清单的信息
def MapBoxModelstage(boxid):
    try:
        check_boxid = list(Box.objects.filter(BoxId=boxid).values())
        check_model = list(ModelStage.objects.filter(ModelStage=check_boxid[0]["ModelStageId"]).values())
        for i in check_model:
            #查找料号相同的sn信息
            search_data = SerialNumberInfo.objects.filter(BoxIdBindingId=check_boxid[0]["Id"]).filter(PartCode=i["PartCode"])
            if i["AlternativePartCode"] !=None:
                search_data=search_data.filter(Q(PartCode=i["PartCode"])|Q(PartCode__in=tuple(i["AlternativePartCode"].split(",")) )).values_list("SerialNumber",flat=True)
            else:
                search_data=search_data.filter(PartCode=i["PartCode"]).values_list("SerialNumber",flat=True)

            i["BrushNumber"]=len(search_data)
            i["BrushSN"] =",".join(search_data)
            if i["BrushNumber"] == i["RequireQuantity"]:
                i["Status"]="OK"
            elif i["BrushNumber"] > i["RequireQuantity"]:
                i["Status"]="Over"
            else:
                i["Status"] = "Open"

        return check_model
    except:
        return []

def MapSNModelstage(SN):
    check_SN = list(SerialNumberInfo.objects.filter(SerialNumber=SN).values())
    if len(check_SN) != 0 and check_SN[0]["BoxIdBindingId"] != None:
        box = list(Box.objects.filter(Id=check_SN[0]["BoxIdBindingId"]).values())
        sns = SerialNumberInfo.objects.filter(BoxIdBindingId=check_SN[0]["BoxIdBindingId"],
                                              PartCode=check_SN[0]["PartCode"]).count()
        sn = SerialNumberInfo.objects.filter(BoxIdBindingId=check_SN[0]["BoxIdBindingId"],
                                             PartCode=check_SN[0]["PartCode"]).values_list("SerialNumber", flat=True)
        data = list(ModelStage.objects.filter(ModelStage=box[0]["ModelStageId"]).filter(
            Q(PartCode=check_SN[0]["PartCode"]) | Q(AlternativePartCode__contains=check_SN[0]["PartCode"])).values())
        if len(data) != 0:
            for i in data:
                i["BrushSN"] = ",".join(sn)
                i["BrushNumber"] = sns
                i["BoxID"] = box[0]["BoxId"]
                if i["RequireQuantity"] == sns:
                    i["Status"] = "OK"
                if i["RequireQuantity"] < sns:
                    i["Status"] = "Over"
            return data
        else:
            Find_no = {}
            Find_no["Id"] = ""
            Find_no["BoxID"] = box[0]["BoxId"]
            Find_no["ModelStage"] = ""
            Find_no["PartCode"] = ""
            Find_no["PartName"] = ""
            Find_no["RequireQuantity"] = ""
            Find_no["BrushNumber"] = 1
            Find_no["BrushSN"] = SN
            Find_no["AlternativePartCode"] = ""
            Find_no["Status"] = "none"
            return data
    else:
        return []
#定义的函数去爬取第三方的数据
#爬取modelstage的信息
def Scrapy_ModelstageId(modelstageId):
    try:
        model_url = "http://10.41.9.13/AEMSWebService/AEMSService.asmx/GetEquipmentList"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36"}
        post_data = {
            "Id": 23,
            "ModelID": modelstageId
        }
        session = requests.session()
        res = session.post(model_url, data=post_data, headers=headers)
        html = etree.HTML(res.content)
        eq = html.xpath("//equipmentlist/equipment")
        data = []
        for i in eq:
            check_link = i.xpath(".//modelpartmandatory/text()")[0]
            if check_link != "0":   #表明是否需要 检查check
                item = {}
                item["ModelStage"] = modelstageId
                item["PartCode"] = str(i.xpath(".//partcode/text()")[0] if len(i.xpath(".//partcode/text()"))>0 else "")
                item["PartName"] = str(i.xpath(".//partname/text()")[0] if len(i.xpath(".//partname/text()"))>0 else "")
                item["RequireQuantity"] = int(i.xpath(".//modelpartquantity/text()")[0] if len(i.xpath(".//modelpartquantity/text()"))>0 else 0)
                item["AlternativeType"] = int(i.xpath(".//modelpartmandatory/text()")[0] if len(i.xpath(".//modelpartmandatory/text()"))>0 else 0)
                data.append(item)
        return data #数据是经过处理的符合python的数据类型的列表包含字典
    except:
        return []
#爬取SN的信息获取数据
def Scrapy_SN(SN):
    try:
        SN_url = "http://10.41.9.13/AEMSWebService/AEMSService.asmx/GetEquipmentInfo"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36"}
        post_data = {
            "Id": 23,
            "BarCode": SN
        }
        session = requests.session()
        res = session.post(SN_url, data=post_data, headers=headers)
        html = etree.HTML(res.content)
        data_sn = []
        item = {}
        item["BarCode"] = str(html.xpath("//equipmentinforesponse//barcode/text()")[0] if len(html.xpath("//equipmentinforesponse//barcode/text()"))>0 else "")
        item["PartCode"] = str(html.xpath("//equipmentinforesponse//partcode/text()")[0] if len(html.xpath("//equipmentinforesponse//partcode/text()"))>0 else "")
        data_sn.append(item)
        return data_sn
    except:
        return []
